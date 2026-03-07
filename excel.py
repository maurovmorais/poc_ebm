import shutil
import os
import openpyxl
import unicodedata

def remover_acentos(arg_strTexto: str) -> str:
    """
    Remove acentos e caracteres especiais de uma string.
    
    Args:
        arg_strTexto (str): Texto original.
        
    Returns:
        str: Texto normalizado sem acentos.
    """
    if not arg_strTexto:
        return ""
    
    # Normaliza para a forma NFD (decomposição de caracteres acentuados)
    var_strNormalizado = unicodedata.normalize('NFD', arg_strTexto)
    # Filtra apenas os caracteres que não são marcas de acentuação (Mn = Mark, Nonspacing)
    return "".join(c for c in var_strNormalizado if unicodedata.category(c) != 'Mn')

def copiar_template_projeto() -> bool:
    """
    Copia o arquivo template.xlsx da raiz do projeto para um novo arquivo 
    chamado extraçao_dados.xlsx na mesma pasta.

    Returns:
        bool: Retorna True se a operação for bem-sucedida, False caso contrário.
    """
    var_strCaminhoRaiz = os.path.dirname(os.path.abspath(__file__))
    var_strArquivoTemplate = os.path.join(var_strCaminhoRaiz, "template.xlsx")
    var_strArquivoDestino = os.path.join(var_strCaminhoRaiz, "extraçao_dados.xlsx")

    var_boolSucesso = False

    try:
        # Verifica se o template existe antes de copiar
        if os.path.exists(var_strArquivoTemplate):
            shutil.copy2(var_strArquivoTemplate, var_strArquivoDestino)
            print(f"Arquivo copiado com sucesso para: {var_strArquivoDestino}")
            var_boolSucesso = True
        else:
            print(f"Erro: Arquivo template não encontrado em {var_strArquivoTemplate}")
            
    except Exception as var_objErro:
        print(f"Ocorreu um erro ao copiar o arquivo: {var_objErro}")

    return var_boolSucesso

def preencher_dados_extracao(arg_listDadosCCV: list, arg_listDadosMin: list) -> bool:
    """
    Preenche o arquivo extraçao_dados.xlsx com as informações extraídas.
    Mapeia as colunas baseadas nos sufixos _CCV e _Min para os respectivos dados.
    Também realiza a comparação automática para as colunas de Status.

    Args:
        arg_listDadosCCV (list): Lista de dicionários com dados do CCV.
        arg_listDadosMin (list): Lista de dicionários com dados do Min.

    Returns:
        bool: True se o preenchimento ocorrer sem erros, False caso contrário.
    """
    var_strCaminhoRaiz = os.path.dirname(os.path.abspath(__file__))
    var_strArquivoExcel = os.path.join(var_strCaminhoRaiz, "extraçao_dados.xlsx")
    var_boolSucesso = False

    if not os.path.exists(var_strArquivoExcel):
        print(f"Erro: Arquivo {var_strArquivoExcel} não encontrado.")
        return False

    try:
        var_objWorkbook = openpyxl.load_workbook(var_strArquivoExcel)
        
        for var_strNomeAba in var_objWorkbook.sheetnames:
            var_objSheet = var_objWorkbook[var_strNomeAba]
            
            # Identifica os cabeçalhos (primeira linha)
            var_listCabecalhos = [cell.value for cell in var_objSheet[1]]
            
            # Localiza a primeira linha vazia (depois do cabeçalho)
            var_intUltimaLinha = var_objSheet.max_row
            # Se a max_row for 1 (apenas cabeçalho), a próxima é a 2.
            # Se a planilha estiver totalmente limpa, o max_row também pode retornar 1.
            var_intPrimeiraLinhaVazia = var_intUltimaLinha + 1 if var_intUltimaLinha > 1 else 2

            # Assume que a quantidade de linhas a preencher é a maior entre as duas listas
            var_intTotalNovasLinhas = max(len(arg_listDadosCCV), len(arg_listDadosMin))
            
            for var_intIndex in range(var_intTotalNovasLinhas):
                var_intLinhaExcel = var_intPrimeiraLinhaVazia + var_intIndex
                var_dictDadosCCV = arg_listDadosCCV[var_intIndex] if var_intIndex < len(arg_listDadosCCV) else {}
                var_dictDadosMin = arg_listDadosMin[var_intIndex] if var_intIndex < len(arg_listDadosMin) else {}

                for var_intColIdx, var_strColuna in enumerate(var_listCabecalhos, 1):
                    if not var_strColuna:
                        continue
                    
                    var_strColunaUpper = var_strColuna.upper()
                    
                    # Preenchimento de dados CCV e Min
                    if var_strColunaUpper.endswith("_CCV"):
                        var_strChave = var_strColuna[:-4] # Remove o _CCV (preservando case do meio se necessário, mas busca será flexível)
                        # Busca insensível a caso no dicionário
                        var_objVal = next((v for k, v in var_dictDadosCCV.items() if k.upper() == var_strChave.upper()), "")
                        var_objSheet.cell(row=var_intLinhaExcel, column=var_intColIdx).value = var_objVal
                        
                    elif var_strColunaUpper.endswith("_MIN"):
                        var_strChave = var_strColuna[:-4] # Remove o _Min
                        var_objVal = next((v for k, v in var_dictDadosMin.items() if k.upper() == var_strChave.upper()), "")
                        var_objSheet.cell(row=var_intLinhaExcel, column=var_intColIdx).value = var_objVal
                    
                # Segunda passada para comparação de Status (após preencher a linha)
                for var_intColIdx, var_strColuna in enumerate(var_listCabecalhos, 1):
                    if not var_strColuna:
                        continue
                        
                    if "STATUS" in var_strColuna.upper():
                        # Lógica aprimorada: Encontrar as colunas _Min e _CCV correspondentes a este Status.
                        # Geralmente o Status está logo após as duas colunas que ele compara.
                        var_strValMin = ""
                        var_strValCCV = ""
                        
                        # Procurar para trás as colunas _Min e _CCV mais próximas
                        for var_intBusca in range(var_intColIdx - 1, 0, -1):
                            var_strColBusca = str(var_listCabecalhos[var_intBusca - 1] or "").upper()
                            if var_strColBusca.endswith("_MIN") and not var_strValMin:
                                var_strValMin = str(var_objSheet.cell(row=var_intLinhaExcel, column=var_intBusca).value or "").strip().upper()
                            elif var_strColBusca.endswith("_CCV") and not var_strValCCV:
                                var_strValCCV = str(var_objSheet.cell(row=var_intLinhaExcel, column=var_intBusca).value or "").strip().upper()
                            
                            if var_strValMin and var_strValCCV:
                                break
                        
                        # Normalização (Remoção de acentos) para comparação
                        var_strValMinNorm = remover_acentos(var_strValMin)
                        var_strValCCVNorm = remover_acentos(var_strValCCV)
                        
                        if var_strValMinNorm == var_strValCCVNorm and var_strValMinNorm != "":
                            var_objSheet.cell(row=var_intLinhaExcel, column=var_intColIdx).value = "OK"
                        else:
                            var_objSheet.cell(row=var_intLinhaExcel, column=var_intColIdx).value = "Divergente"

        var_objWorkbook.save(var_strArquivoExcel)
        print("Dados preenchidos com sucesso no Excel.")
        var_boolSucesso = True

    except Exception as var_objErro:
        print(f"Erro ao preencher o Excel: {var_objErro}")
    
    return var_boolSucesso
